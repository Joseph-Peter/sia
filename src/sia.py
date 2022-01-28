from pathlib import Path
#import os
import spacy
import streamlit as st
import re
from collections import defaultdict
import pandas as pd
import numpy as np
#import xlwings as xw
from openpyxl import load_workbook

# Monkey patch to offset bug when pptx is used on Windows and Python 3.10
# import collections
# import collections.abc
# c = collections
# c.abc = collections.abc
# from pptx import Presentation
import tika
from tika import parser
tika.initVM()
# Tika hides the complexity of different file formats and parsing libraries while providing a simple and powerful
# #mechanism for client applications to extract structured text content and metadata from all sorts of documents.
from io import FileIO
from spacy.matcher import PhraseMatcher

TYPES_OF_KEYWORDS = ['Technical',
                     'Business',
                     'Theme',
                     'User Input',
                    ]
LIST_OF_USEFUL_COLS = [
        'Resource Personnel #',
        'Resource Full Name',
        'Resource Management Level',
        'Mudano Role Override',
     #  'Resource Availability Detailed Status',
        'Resource First Available Date',
        'ACN location',
        'Community/Capability',
     #  'Tags',
        'Source',
        'Gender',
        'Req',
     #  '1st check in date',
     #  '2nd check in date',
    ]

class TwoWayMappingStore:
    def __init__(self):
        self.def_dict = defaultdict(list)
        self.normal_dict = dict()

    def set(self, key, value):
        if type(value) == str:
            self.def_dict[key].append(value)
            self.normal_dict[value] = key
        elif type(value) == list:
            for s in value:
                self.set(key, s)

    def merge(self, two_way_mapping_store_2):
        self.def_dict.update(two_way_mapping_store_2.def_dict)
        self.normal_dict.update(two_way_mapping_store_2.normal_dict)

    def get_forward(self, key):
        return self.def_dict[key]

    def get_reverse(self, value):
        return self.normal_dict[value]

    def print(self):
        print(self.def_dict)


class KeywordHolder:
    def __init__(self):
        self.individual_keywords = []
        self.categorical_keywords = []
        self.mappings_from_categorical = TwoWayMappingStore()

    def add_individual(self, k):
        self.individual_keywords.append(k)

    def add_categorical(self, list_of_keywords):
        category = list_of_keywords[0]
        self.categorical_keywords.append(category)
        self.individual_keywords.extend(list_of_keywords[1:])
        self.mappings_from_categorical.set(category, list_of_keywords[1:])

    @staticmethod
    def _print_dict(dict_to_be_printed):
        for count, (k, v) in enumerate(dict_to_be_printed.items()):
            print(f"{count}. {k} -> {v}")

    def return_list_of_all_keywords(self):
        return self.categorical_keywords + self.individual_keywords

    def return_list_of_individual_keywords(self):
        return  self.individual_keywords

    def return_list_of_categorical_keywords(self):
        return self.categorical_keywords

    def print(self):
        print("Categories :")
        for count, cat in enumerate(self.categorical_keywords):
            print(f"{count+1}. {cat}")
        print("Individual keywords :")
        for count, ind in enumerate(self.individual_keywords):
            print(f"{count + 1}. {ind}")
        print("Mappings to Categorical:")
        print('\n')
        self.mappings_from_categorical.print()


class KeywordManager:
    def __init__(self, types_of_keywords=None):
        self.types_of_keywords = types_of_keywords
        self.count_of_all_keywords = 0
        self.mappings_from_categorical = TwoWayMappingStore()
        self.dict_of_keyword_holders = dict()
        for t in self.types_of_keywords:
            self.dict_of_keyword_holders[t] = KeywordHolder()

    @property
    def set_of_all_keywords(self):
        s = set()
        for t in self.types_of_keywords:
            s.update(self.dict_of_keyword_holders[t].return_list_of_all_keywords())
        return s

    @property
    def set_of_individual_keywords(self):
        s = set()
        for t in self.types_of_keywords:
            s.update(self.dict_of_keyword_holders[t].return_list_of_individual_keywords())
        return s

    @property
    def set_of_categorical_keywords(self):
        s = set()
        for t in self.types_of_keywords:
            s.update(self.dict_of_keyword_holders[t].return_list_of_categorical_keywords())
        return s

    def check_all_keywords_unique(self):
        if len(self.set_of_all_keywords) == self.count_of_all_keywords:
            return True
        else:
            return False

    def add_line_of_keywords(self, type_of_keywords, line):
        l = len(line)
        if l == 0:
            return
        elif l == 1:
            self.dict_of_keyword_holders[type_of_keywords].add_individual(line[0])
            self.count_of_all_keywords += 1
        else:
            self.dict_of_keyword_holders[type_of_keywords].add_categorical(line)
            self.count_of_all_keywords += l

    def add_type_of_keywords(self, type_of_keywords, list_of_list_of_keywords):
        for line in list_of_list_of_keywords:
            self.add_line_of_keywords(type_of_keywords=type_of_keywords,
                                      line=line,
                                      )

    def merge_all_mappings(self):
        for t in self.types_of_keywords:
            self.mappings_from_categorical.merge(self.dict_of_keyword_holders[t].mappings_from_categorical)

    def print(self):
        for keyword_type in self.types_of_keywords:
            print(keyword_type)
            self.dict_of_keyword_holders[keyword_type].print()

def get_email_addresses(s):
    return re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', s)


# method for searching keyword from the text
def search_for_keyword(keyword_to_be_searched, doc_obj, nlp):
    phrase_matcher = PhraseMatcher(vocab=nlp.vocab, attr='LOWER')
    phrase_list = [nlp(keyword_to_be_searched.lower())]
    phrase_matcher.add("Text Extractor", phrase_list)
    # words_found = []
    # contexts_found = []
    # for match_id, start, end in matched_items:
    #     words_found.append(doc_obj[start: end])
    #     try:
    #         contexts_found.append(str(doc_obj[start - 20: end + 20]))
    #     except IndexError:
    #         contexts_found.append("")
    # contexts_found = [context.split(None, 1)[1] for context in contexts_found]
    # contexts_found = [' '.join(context.split()[:-1]) for context in contexts_found]
    # return words_found, contexts_found

    matched_items = phrase_matcher(doc_obj)
    return len(matched_items)

# Reading pptx file directly from url is not trivially done
# SO question on this https://stackoverflow.com/questions/43589798/read-pptx-file-content-from-a-url# is unanswered
# import urllib  # the lib that handles the url stuff
# target_url_root = "ts.accenture.com"
# target_url = "https://ts.accenture.com/:p:/r/sites/Mudano_Business/BusinessSupport/Mudano%20Format%20CVs/Client%20Ready%20CVs/ARC%20-%20Mudano%20CV%20Anurag%20Kataria%20Mar%2020.pptx?d=wd2258dda78b840d28339ab150dabc706&csf=1&web=1&e=QlpPm5"
# file = urllib.request.urlopen(target_url)

def scan_for_input_list(st_in, data_folder, input_all_individual_list):
    nlp = spacy.load('en_core_web_sm')
    data = defaultdict(list)
    for file in Path(data_folder).glob('*.pptx'):
        parsed = parser.from_file(FileIO(Path(file).absolute(), "rb"))
        text = parsed["content"]
        # text = os.linesep.join([s for s in text.splitlines() if s is not None])
        if text:
            emails = get_email_addresses(text)
        else:
            emails = []
        if len(emails) == 0:
            st_in.warning(f"Warning: No email found in {file.name}")
        elif len(emails) > 1:
            st_in.warning(f"Warning: More than one email found in {file.name}")
        else:
            doc = nlp(text)
            data['Filename'].append(file.name)
            data['Email'].append(emails[0])
            for individual_keyword in input_all_individual_list:
                no_of_occurrences = search_for_keyword(individual_keyword, doc, nlp)
                if not no_of_occurrences:
                    no_of_occurrences = 0
                data[individual_keyword].append(no_of_occurrences)
    df = pd.DataFrame(data)
    df['Name from Email'] = df['Email'].apply(lambda x: " ".join(x.split('@')[0].split('.')).title())
    return df

def run_a_query(st_in, input_individual_list, input_categorical_list, scan_again, two_way_dict_mapping):
    #data_folder = Path("C:/Users/joseph.peter/OneDrive - Accenture/Documents/dev/delete_later/Client Ready CVs/")
    #data_folder = Path("C:/Users/joseph.peter/Documents/Client Ready CVs")
    #data_folder = Path("https://ts.accenture.com/:f:/r/sites/StabilisationTest/Shared%20Documents/myCV?csf=1&web=1&e=Zo1w0H")
    data_folder = Path("D:/Code/Python/SIA/myCV/")
    excel_path = 'D:/Code/Python/SIA/Mudano Resourcing 2022.xlsx'

    input_all_individual_list = input_individual_list.copy()
    for categorical_input in input_categorical_list:
        input_all_individual_list.extend(two_way_dict_mapping.get_forward(categorical_input))
    input_all_individual_list = list(set(input_all_individual_list))

    if scan_again:
        df = scan_for_input_list(st_in=st_in,
                                 data_folder=data_folder,
                                 input_all_individual_list=input_all_individual_list,
                                 )
        df_temp = df.select_dtypes(include='number')
        if len(df_temp.columns) > 3:
            df_temp = df_temp.apply(lambda x: pd.Series(x.sort_values(ascending=False).iloc[:3].index,
                               index=['top1', 'top2', 'top3']),
                               axis=1).reset_index()
            df_temp['Skill_summary'] = df_temp['top1'] + '/' + df_temp['top2'] + '/' + df_temp['top3']
            df['Skill_summary'] = df_temp['Skill_summary'].copy()

        df.to_csv('D:/Code/Python/SIA/cv_analysis_results.csv', index=False)
        #print('written file')
        df = pd.read_csv('D:/Code/Python/SIA/cv_analysis_results.csv')
        missing_cols = [column for column in input_all_individual_list if column not in df.columns]
        if missing_cols:
            missing_col_df = scan_for_input_list(st_in=st_in,
                                                 data_folder=data_folder,
                                                 input_all_individual_list=missing_cols,
                                                )
            df = df.merge(missing_col_df,
                          how='outer',
                          left_on=['Name from Email', 'Email', 'Filename'],
                          right_on=['Name from Email', 'Email', 'Filename'],
                          suffixes=(None, None),
                          )
            df.to_csv(
                'D:/Code/Python/SIA/cv_analysis_results.csv',
                index=False)

        df = df.drop([column for column in df.columns if column not in (['Name from Email', 'Email', 'Filename', 'Skill_summary'] + input_all_individual_list)], axis=1)

    q = ""
    first = True
    for key in input_individual_list:
        if ' ' in key:
            n_key = '`' + key +  '`'
        else:
            n_key = key
        if not first:
            q += ' and '
        q += '(' + n_key + ' > 0) '
        if first:
            first = False

    for key in input_categorical_list:
        if not first:
            q += ' and '
        first_categorical = True
        q += '('
        l = two_way_dict_mapping.get_forward(key)
        for individual_key in l:
            if ' ' in individual_key:
                n_key = '`' + individual_key + '`'
            else:
                n_key = individual_key
            if not first_categorical:
                q += ' or '
            q += '(' + n_key + ' > 0) '
            if first_categorical:
                first_categorical = False
        q += ')'

    result = df.query(q)
    # prs = Presentation(file)
    # text = ""
    # for slide in prs.slides:
    #     for shape in slide.shapes:
    #         if not shape.has_text_frame:
    #             continue
    #         for paragraph in shape.text_frame.paragraphs:
    #             for run in paragraph.runs:
    #                 text += run.text + "\n"
    #
    #
    # for slide in prs.slides:
    #     for shapes in slide.shapes:
    #         if shapes.has_text_frame:
    #             text += shapes.text + "\n"

    resource_info = pd.read_excel(excel_path, sheet_name='supply_new', header=0)
    list_of_useful_cols = LIST_OF_USEFUL_COLS
    resource_info = resource_info[list_of_useful_cols]
    merged_df = result.merge(resource_info,
                             how='left',
                             left_on='Name from Email',
                             right_on='Resource Full Name',
                             suffixes=(None, None))

    first_column = merged_df.pop('Name from Email')
    merged_df.insert(0, 'Name from Email', first_column)
    last_column = merged_df.pop('Filename')
    merged_df.insert(13, 'Filename', last_column)
    merged_df.sort_values(by=['Resource First Available Date',
                              'Name from Email',
                              ],
                          inplace=True)
    merged_df.index = np.arange(1, len(merged_df) + 1)
    return merged_df

# MAIN PROGRAM

#st.title('Mudano Skills Inventory Analysis')
top_title = '<p style="font-family:Courier; color:rgb(123, 39, 225); font-size: 50px;"> <b>Mudano Skills Inventory Analysis </b></p>'
st.markdown(top_title, unsafe_allow_html=True)
st.markdown("""
This app gives the user a list of resources that are available at Mudano for a user-specified set of skills.
The results are presented after sorting so that the earliest available resource will come at the top.
""")

side_title = '<p style="font-family:Courier; color:rgb(123, 39, 225); font-size: 20px;"> <b>Query Input (Skills to search for)</b></p>'
st.sidebar.markdown(side_title, unsafe_allow_html=True)

wb = load_workbook(filename = "D:/Code/Python/SIA/keywords.xlsx")
km = KeywordManager(types_of_keywords=TYPES_OF_KEYWORDS)

for i, sheet in enumerate(wb):
    list_of_lists_for_sheet = []
    for row in range(1, 100):
        list_for_row = []
        for col in range(1, 10):
            c = sheet.cell(row=row, column=col).value
            if c is None:
                break
            else:
                list_for_row.append(c)
        if not list_for_row:
            break
        list_of_lists_for_sheet.append(list_for_row)
    km.add_type_of_keywords(type_of_keywords=wb.sheetnames[i],
                            list_of_list_of_keywords=list_of_lists_for_sheet,
                            )

#km.print()
if not km.check_all_keywords_unique():
    st.error(f"Error: Some search keyword(s) are included more than once")
else:
    technical_keyword_list = km.dict_of_keyword_holders[TYPES_OF_KEYWORDS[0]].return_list_of_all_keywords()
    selected_technical_keywords = st.sidebar.multiselect('Technical skills', technical_keyword_list, [])

    vertical_expertise_list = km.dict_of_keyword_holders[TYPES_OF_KEYWORDS[1]].return_list_of_all_keywords()
    selected_vertical_keywords = st.sidebar.multiselect('Industry / Vertical expertise', vertical_expertise_list, [])

    themes_list = km.dict_of_keyword_holders[TYPES_OF_KEYWORDS[2]].return_list_of_all_keywords()
    selected_theme_keywords = st.sidebar.multiselect('Project themes', themes_list, [])

    user_input = st.sidebar.text_area("Enter your own keywords in separate lines below. A group of keywords can be entered on the same line separated by commas.", value="", height=100)
    user_input = user_input.splitlines()
    user_input_list_of_lists = [line.split(',') for line in user_input]
    km.add_type_of_keywords(type_of_keywords=TYPES_OF_KEYWORDS[3],
                            list_of_list_of_keywords=user_input_list_of_lists,
                            )
    if not km.check_all_keywords_unique():
        st.error(f"Error: Some search keyword(s) are included more than once")
    else:
        km.merge_all_mappings()
        scan_again_input = st.sidebar.checkbox('Scan all CVs again')
        save_all_input = st.sidebar.checkbox('Save results of all predefined keywords')
        if save_all_input:
            scan_again_input = True

        if save_all_input:
            selected_keywords = technical_keyword_list + vertical_expertise_list + themes_list
            #Not including user input keywords here
        else:
            selected_keywords = selected_technical_keywords + selected_vertical_keywords + selected_theme_keywords + user_input

        merged_dict = km.mappings_from_categorical
        set_of_selected_individual_keywords = set()
        set_of_selected_categorical_keywords = set()
        for keyword in selected_keywords:
            if keyword in merged_dict.def_dict.keys():
                set_of_selected_categorical_keywords.add(keyword)
            else:
                set_of_selected_individual_keywords.add(keyword)
        m = st.markdown("""
        <style>
        div.stButton > button:first-child {
            background-color: rgb(123, 39, 225);
        }
        </style>""", unsafe_allow_html=True)
        if st.sidebar.button('Run query'):
            if not selected_keywords:
                st.warning("Warning: No keywords selected")
            else:
                results = run_a_query(st_in=st,
                                      input_individual_list=list(set_of_selected_individual_keywords),
                                      input_categorical_list=list(set_of_selected_categorical_keywords),
                                      scan_again=scan_again_input,
                                      two_way_dict_mapping=km.mappings_from_categorical,
                                     )

                for col in set_of_selected_categorical_keywords:
                    results[col] = results[merged_dict.get_forward(col)].sum(axis=1)

                cols_needed = set_of_selected_individual_keywords\
                              .union(set_of_selected_categorical_keywords)\
                              .union(set(LIST_OF_USEFUL_COLS)) \
                              .union(set(['Name from Email', 'Email', 'Skill_summary']))
                cols_to_be_dropped = [col for col in results if col not in cols_needed]
                results = results.drop(columns=cols_to_be_dropped)
                #st.header('Available resources for last query')
                if save_all_input:
                    second_title = '<p style="font-family:Courier; color:Purple; font-size: 30px;"> <b>All pre-defined keywords scanned and saved </b></p>'
                    st.markdown(second_title, unsafe_allow_html=True)
                else:
                    second_title = '<p style="font-family:Courier; color:Purple; font-size: 30px;"> <b>Available resources for last query </b></p>'
                    st.markdown(second_title, unsafe_allow_html=True)
                    st.write(f'{len(results)} resources found')

                    #Reorder columns
                    column_display_order = list()
                    column_display_order.append('Name from Email')
                    column_display_order.extend(list(set_of_selected_categorical_keywords))
                    column_display_order.extend(list(set_of_selected_individual_keywords))
                    if 'Skill_summary' in results:
                        column_display_order.append('Skill_summary')
                    column_display_order.append('Resource First Available Date')
                    column_display_order.append('Email')

                    for col in results:
                        if col not in column_display_order:
                            column_display_order.append(col)

                    results = results[column_display_order]

                    st.write(results.style
                                 .set_table_styles([{'selector' : 'th', 'props' : [('border','5px solid green')]}]) \
                                 .set_properties(**{'background-color': 'lavender', 'color': 'purple','border-color': 'white'}),
                                 width=2048,
                                 height=768)